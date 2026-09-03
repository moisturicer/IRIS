from io import BytesIO

from rest_framework import viewsets, mixins, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
import jwt

from core.permissions import IsOwnerOrStaff, IsReviewer, IsStaff, IsAdmin, IsRDCO
from .download_service import file_response_for_record
from .download_tokens import make_download_token, verify_download_token
from .models import Record, DownloadRequest, DeleteRequest
from .serializers import (
    RecordListSerializer,
    RecordDetailSerializer,
    RecordWriteSerializer,
    DownloadRequestSerializer,
    DeleteRequestSerializer,
)
from .filters import RecordFilter
from .services import soft_delete_record, parse_excel_import
from .download_tokens import make_download_token, verify_download_token
from .download_service import file_response_for_record
from apps.notifications.services import (
    notify_new_record,
    notify_download_request,
    notify_download_reviewed,
    notify_delete_approved,
    notify_delete_declined,
)
from apps.audit.services import create_audit_event


class RecordViewSet(viewsets.ModelViewSet):
    """
    GET    /records/           -- published records (anyone authenticated)
    POST   /records/           -- create (student/adviser+)
    GET    /records/<id>/      -- detail
    PATCH  /records/<id>/      -- update (owner only, not yet approved)
    DELETE /records/<id>/      -- soft delete / create DeleteRequest
    """
    # Restrict pk lookups to integers so that sub-paths like record-types/,
    # classifications/, etc. are not consumed by this ViewSet's detail route.
    lookup_value_regex = r'\d+'
    filter_backends  = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class  = RecordFilter
    search_fields    = ["title", "abstract", "authors__name"]
    ordering_fields  = ["created_at", "year_accomplished", "title", "access_count"]
    ordering         = ["-created_at"]

    def get_queryset(self):
        # Public list shows published research, approved (ongoing) and completed proposals
        if self.action == "list":
            # distinct=True: the college/department filters join through owners.
            return Record.objects.publicly_visible().annotate(
                file_count=Count("files", distinct=True)
            ).select_related(
                "classification", "psced", "record_type", "adviser"
            ).prefetch_related("owners__user", "authors")
        return Record.objects.select_related(
            "classification", "psced", "record_type", "adviser"
        ).prefetch_related("owners__user", "authors")

    def get_serializer_class(self):
        if self.action == "list":
            return RecordListSerializer
        if self.action in ("create", "update", "partial_update"):
            return RecordWriteSerializer
        return RecordDetailSerializer

    def get_permissions(self):
        if self.action in ("update", "partial_update", "destroy"):
            return [IsAuthenticated(), IsOwnerOrStaff()]
        if self.action == "complete":
            return [IsAuthenticated(), IsRDCO()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        from .models import RecordOwner
        record = serializer.save(added_by=self.request.user, pipeline_status="draft")
        # Add the creator as the primary owner automatically
        RecordOwner.objects.create(record=record, user=self.request.user, is_primary=True)
        # Record starts as draft — notification fires only when the owner calls /submit/

    def perform_destroy(self, instance):
        # Publicly visible records go through delete request flow
        if instance.pipeline_status in ("published", "approved", "completed"):
            DeleteRequest.objects.create(
                record=instance,
                requested_by=self.request.user,
                previous_pipeline_status=instance.pipeline_status,
            )
            instance.pipeline_status = "pending_delete"
            instance.save(update_fields=["pipeline_status"])
        else:
            soft_delete_record(instance, deleted_by=self.request.user)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsOwnerOrStaff])
    def submit(self, request, pk=None):
        """
        POST /records/<id>/submit/

        Transition a new draft record into the pipeline.
        Routing depends on record type:
          Proposal        -> adviser_review   (adviser notified)
          Thesis/Research -> rdco_intake      (RDCO notified)
          Project         -> rdco_intake      (RDCO notified)

        Rules:
          - Record must be in 'draft' status.
          - Proposal: adviser must be assigned.
          - record_type must be set.

        Also handles resubmission after revision ('declined' → owner fixes and resubmits).
        'rejected' is the terminal state — rejected records cannot be resubmitted.
        """
        record = self.get_object()  # enforces IsOwnerOrStaff object permission

        if record.pipeline_status not in ("draft", "declined"):
            return Response(
                {"detail": f"Record is in '{record.pipeline_status}' status and cannot be submitted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not record.record_type:
            return Response(
                {"detail": "A record type must be selected before the record can be submitted."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rt_name = record.record_type.name  # "Proposal" | "Thesis/Research" | "Project"

        if rt_name == "Proposal":
            if not record.adviser:
                return Response(
                    {"detail": "An adviser must be assigned before a Proposal can be submitted."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            first_status = "adviser_review"
        else:
            first_status = "rdco_intake"

        record.pipeline_status = first_status
        record.save(update_fields=["pipeline_status", "updated_at"])

        # Notify the correct party — never raises (wrapped inside the service)
        notify_new_record(record, submitted_by=request.user)

        stage_label = "adviser" if rt_name == "Proposal" else "RDCO"
        return Response(
            {"detail": f"Record submitted successfully. The {stage_label} has been notified."},
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["get"])
    def similar(self, request, pk=None):
        """
        GET /records/<id>/similar/ — related institutional works.

        Reuses the Ask IRIS retrieval service, so "similar" means the same
        ranking users get from search, over the same visibility predicate.
        """
        from apps.ai.services.retrieval import search_records

        record = self.get_object()
        seed = f"{record.title} {record.abstract or ''}".strip()
        matches = search_records(seed, top_k=3, exclude_id=record.id)
        return Response({"results": [s.as_dict() for s in matches]})

    @action(detail=True, methods=["post"])
    def increment_access(self, request, pk=None):
        """POST /records/<id>/increment_access/"""
        record = self.get_object()
        Record.objects.filter(pk=record.pk).update(access_count=record.access_count + 1)
        create_audit_event("ACCESS", request.user, record=record)
        return Response({"access_count": record.access_count + 1})

    @action(detail=True, methods=["patch"], permission_classes=[IsAuthenticated, IsStaff])
    def tags(self, request, pk=None):
        """
        PATCH /records/<id>/tags/
        Staff (KTTO, RDCO, ITSO, IERC) may update IP classification
        flags and the structured ip_type label on any record.

        Boolean fields (all optional):
          { "is_ip": true, "for_commercialization": false, "community_extension": false }

        IP type field (FR-M5-05):
          { "ip_type": "patent" | "copyright" | "trade_secret" | "utility_model" | "" }

        Passing ip_type="" clears the classification.
        """
        from .models import Record as RecordModel

        record = self.get_object()

        BOOL_FIELDS   = {"is_ip", "for_commercialization", "community_extension"}
        VALID_IP_TYPES = {"patent", "copyright", "trade_secret", "utility_model", ""}
        updates: dict = {}

        for field in BOOL_FIELDS:
            if field in request.data:
                value = request.data[field]
                if not isinstance(value, bool):
                    return Response(
                        {"detail": f"'{field}' must be a boolean (true or false)."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                updates[field] = value

        if "ip_type" in request.data:
            ip_type_value = request.data["ip_type"]
            if ip_type_value not in VALID_IP_TYPES:
                return Response(
                    {
                        "detail": (
                            f"'{ip_type_value}' is not a valid ip_type. "
                            "Choose from: patent, copyright, trade_secret, utility_model "
                            "(or '' to clear)."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            updates["ip_type"] = ip_type_value

        if not updates:
            return Response(
                {
                    "detail": (
                        "No valid tag fields provided. "
                        "Use: is_ip, for_commercialization, community_extension, ip_type."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        for field, value in updates.items():
            setattr(record, field, value)
        record.save(update_fields=list(updates.keys()) + ["updated_at"])

        create_audit_event(
            "ACCESS", request.user, record=record,
            metadata={"action": "tags_updated", "changes": updates},
        )

        return Response(RecordDetailSerializer(record, context={"request": request}).data)

    @action(detail=True, methods=["post"])
    def complete(self, request, pk=None):
        """
        POST /records/<id>/complete/
        RDCO marks an approved Proposal as completed (research finished).
        The record remains publicly visible.
        Permission enforced by get_permissions() → IsRDCO.
        """
        from apps.notifications.services import notify_proposal_completed

        record = self.get_object()

        if record.pipeline_status != "approved":
            return Response(
                {"detail": f"Only approved proposals can be marked as completed (current status: '{record.pipeline_status}')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rt_name = record.record_type.name if record.record_type else ""
        if rt_name != "Proposal":
            return Response(
                {"detail": "Only Proposal records can be marked as completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        record.pipeline_status = "completed"
        record.save(update_fields=["pipeline_status", "updated_at"])

        notify_proposal_completed(record, marked_by=request.user)

        return Response({"detail": "Proposal marked as completed."}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"])
    def mine(self, request):
        """GET /records/mine/ -- records the current user owns.
        Optional: ?pipeline_status=published,approved  (comma-separated)
        """
        qs = Record.objects.filter(owners__user=request.user).distinct()
        statuses = request.query_params.get("pipeline_status")
        if statuses:
            qs = qs.filter(pipeline_status__in=[s.strip() for s in statuses.split(",")])
        serializer = RecordListSerializer(qs, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=False, methods=["post"], permission_classes=[IsAuthenticated, IsStaff])
    def import_excel(self, request):
        """
        POST /records/import_excel/
        Accepts a multipart upload with key "file" (.xls or .xlsx).
        Creates published records from each valid row and returns a plaintext log.
        Legacy imports bypass the review pipeline — staff is the implicit reviewer.
        """
        from .models import Author
        from .models import RecordType, Classification, PSCEDClassification

        file = request.FILES.get("file")
        if not file:
            return Response(
                {"detail": "No file uploaded.", "log": "Error: no file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows, parse_errors = parse_excel_import(file)
        except Exception as exc:
            return Response(
                {"detail": str(exc), "log": f"Parse error: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        log_lines: list[str] = []
        for err in parse_errors:
            log_lines.append(f"⚠  {err}")

        created = 0
        failed  = len(parse_errors)

        for row in rows:
            try:
                record_type = RecordType.objects.filter(
                    name__iexact=row["record_type_name"]
                ).first()

                classification = (
                    Classification.objects.filter(name__iexact=row["classification_name"]).first()
                    if row["classification_name"] else None
                )
                psced = (
                    PSCEDClassification.objects.filter(name__iexact=row["psced_name"]).first()
                    if row["psced_name"] else None
                )

                record = Record.objects.create(
                    title                = row["title"],
                    abstract             = row["abstract"],
                    year_accomplished    = row["year_accomplished"],
                    year_completed       = row["year_completed"],
                    record_type          = record_type,
                    classification       = classification,
                    psced                = psced,
                    is_ip                = row["is_ip"],
                    for_commercialization= row["for_commercialization"],
                    community_extension  = row["community_extension"],
                    added_by             = request.user,
                    pipeline_status      = "published",
                )

                from .models import RecordOwner
                RecordOwner.objects.create(record=record, user=request.user, is_primary=True)

                for author_name in row["authors"]:
                    Author.objects.create(record=record, name=author_name)

                log_lines.append(f"✓  Created: {record.title[:70]}")
                created += 1

            except Exception as exc:
                log_lines.append(f"✗  Failed:  {row['title'][:70]} — {exc}")
                failed += 1

        log_lines.append(f"\n{'─' * 50}")
        log_lines.append(f"Done — {created} record(s) created, {failed} skipped.")
        return Response({
            "log":     "\n".join(log_lines),
            "created": created,
            "skipped": failed,
        })

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated, IsStaff])
    def download_template(self, request):
        """
        GET /records/download_template/
        Returns a pre-formatted .xlsx import template with a sample data row
        and an END OF RECORDS sentinel so users know where to stop.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response(
                {"detail": "openpyxl is not installed on the server."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"

        HEADERS = [
            "Title",
            "Abstract",
            "Year Accomplished",
            "Year Completed",
            "Record Type",
            "Classification",
            "PSCED Classification",
            "Is IP",
            "For Commercialization",
            "Community Extension",
            "Authors",
        ]
        WIDTHS = [55, 65, 18, 16, 22, 28, 36, 10, 22, 22, 45]

        # ---- Header row styling --------------------------------------------
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill(start_color="6B0F12", end_color="6B0F12", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin",  color="CCCCCC"),
        )

        for col_idx, (header, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 30

        # ---- Example data row ----------------------------------------------
        ws.append([
            "Blockchain-Based Verification of Academic Credentials",
            "This study proposes a tamper-proof academic credential system using distributed ledger technology.",
            2024,
            2024,
            "Thesis/Research",
            "Applied Sciences",
            "Information and Communication Technologies",
            "FALSE",
            "FALSE",
            "FALSE",
            "Juan Dela Cruz, Maria Santos",
        ])

        # ---- Notes row (light grey) ----------------------------------------
        notes_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        notes_font = Font(italic=True, size=9, color="856404")
        note_row   = [
            "Required",
            "Optional",
            "e.g. 2024",
            "e.g. 2024",
            "Proposal / Thesis/Research / Project",
            "Exact name from the system",
            "Exact name from the system",
            "TRUE or FALSE",
            "TRUE or FALSE",
            "TRUE or FALSE",
            "Comma-separated names",
        ]
        ws.append(note_row)
        note_row_idx = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=note_row_idx, column=col_idx)
            cell.fill = notes_fill
            cell.font = notes_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ---- Sentinel row --------------------------------------------------
        ws.append(["END OF RECORDS"])
        sentinel_font = Font(bold=True, color="FFFFFF")
        sentinel_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        sentinel_cell = ws.cell(row=ws.max_row, column=1)
        sentinel_cell.font = sentinel_font
        sentinel_cell.fill = sentinel_fill

        # ---- Freeze header row ---------------------------------------------
        ws.freeze_panes = "A2"

        # ---- Serialize and return ------------------------------------------
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="iris_import_template.xlsx"'
        return response


class MyRecordsViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    GET /records/mine/          -- list records I own (with pipeline_status)
    GET /records/mine/<id>/     -- detail with full review history
    """
    serializer_class = RecordDetailSerializer

    def get_queryset(self):
        return Record.objects.filter(owners__user=self.request.user).distinct()


class DownloadRequestViewSet(viewsets.ModelViewSet):
    """
    GET    /download-requests/                    -- staff: all requests
    POST   /download-requests/                    -- authenticated user requests download
    POST   /download-requests/<id>/approve/       -- staff: approve, email requester
    POST   /download-requests/<id>/decline/       -- staff: decline, in-app notify requester
    """
    serializer_class = DownloadRequestSerializer
    queryset         = DownloadRequest.objects.select_related("record", "requested_by")
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_permissions(self):
        if self.action == "create":
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsStaff()]

    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs.order_by("-created_at")

    def perform_create(self, serializer):
        record = serializer.validated_data["record"]
        user   = self.request.user
        if DownloadRequest.objects.filter(
            record=record, requested_by=user, status="pending"
        ).exists():
            from rest_framework.exceptions import ValidationError
            raise ValidationError(
                {"record": ["You already have a pending download request for this record."]}
            )
        dr = serializer.save(requested_by=user)
        notify_download_request(dr.record, requested_by=user)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        action = request.data.get("action")
        if action not in ("approve", "decline"):
            return Response(
                {"detail": "Provide action: 'approve' or 'decline'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if instance.status != "pending":
            return Response(
                {"detail": "This request has already been reviewed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        instance.reviewed_by = request.user
        instance.reviewed_at = timezone.now()
        instance.status = "approved" if action == "approve" else "declined"
        instance.save(update_fields=["status", "reviewed_by", "reviewed_at"])

        data = self.get_serializer(instance).data
        if action == "approve":
            token = make_download_token(
                download_request_id=instance.id,
                record_id=instance.record_id,
                user_id=instance.requested_by_id,
            )
            data["download_url"] = f"{settings.FRONTEND_URL.rstrip('/')}/download?token={token}"
        return Response(data)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsStaff])
    def approve(self, request, pk=None):
        """POST /download-requests/<id>/approve/ — set approved, notify requester with email."""
        dr = self.get_object()
        if dr.status != "pending":
            return Response(
                {"detail": f"Request is already '{dr.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        dr.status      = "approved"
        dr.reviewed_by = request.user
        dr.reviewed_at = timezone.now()
        dr.save(update_fields=["status", "reviewed_by", "reviewed_at"])
        notify_download_reviewed(dr, reviewed_by=request.user, approved=True)
        return Response({"detail": "Download request approved. The requester has been notified."})

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsStaff])
    def decline(self, request, pk=None):
        """POST /download-requests/<id>/decline/ — set declined, notify requester in-app."""
        dr = self.get_object()
        if dr.status != "pending":
            return Response(
                {"detail": f"Request is already '{dr.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        dr.status      = "declined"
        dr.reviewed_by = request.user
        dr.reviewed_at = timezone.now()
        dr.save(update_fields=["status", "reviewed_by", "reviewed_at"])
        notify_download_reviewed(dr, reviewed_by=request.user, approved=False)
        return Response({"detail": "Download request declined. The requester has been notified."})


class DownloadRedeemView(APIView):
    """
    GET /records/download/?token=<jwt>
    Redeem an approved download JWT and stream the record PDF.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        token = request.query_params.get("token")
        if not token:
            return Response({"detail": "token is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            claims = verify_download_token(token)
        except jwt.ExpiredSignatureError:
            return Response({"detail": "Download link has expired."}, status=status.HTTP_403_FORBIDDEN)
        except jwt.InvalidTokenError:
            return Response({"detail": "Invalid download link."}, status=status.HTTP_403_FORBIDDEN)

        try:
            dl_request = DownloadRequest.objects.select_related("record").get(
                pk=claims["drid"], status="approved"
            )
        except DownloadRequest.DoesNotExist:
            return Response({"detail": "Download request not found or not approved."}, status=404)

        if dl_request.record_id != claims["rid"] or dl_request.requested_by_id != claims["uid"]:
            return Response({"detail": "Invalid download link."}, status=status.HTTP_403_FORBIDDEN)

        response = file_response_for_record(dl_request.record)
        if not response:
            return Response(
                {"detail": "No downloadable file is available for this record."},
                status=status.HTTP_404_NOT_FOUND,
            )
        # TODO(SRS): apply per-user watermark (email, date) before streaming when required
        return response


class DeleteRequestViewSet(viewsets.ModelViewSet):
    """
    GET    /delete-requests/                  -- staff: all requests
    POST   /delete-requests/                  -- authenticated user requests deletion
    POST   /delete-requests/<id>/approve/     -- admin: soft-delete record, notify owner
    """
    serializer_class = DeleteRequestSerializer
    queryset         = DeleteRequest.objects.select_related("record", "requested_by")

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated(), IsStaff()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsRDCO])
    def approve(self, request, pk=None):
        """
        POST /delete-requests/<id>/approve/
        Marks the request approved, soft-deletes the record, and notifies the requester.
        """
        from django.utils import timezone
        dr = self.get_object()
        if dr.status != "pending":
            return Response(
                {"detail": f"Request is already '{dr.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        dr.status      = "approved"
        dr.reviewed_by = request.user
        dr.reviewed_at = timezone.now()
        dr.save(update_fields=["status", "reviewed_by", "reviewed_at"])
        soft_delete_record(dr.record, deleted_by=request.user)
        notify_delete_approved(dr, reviewed_by=request.user)
        return Response({"detail": "Delete request approved. The record has been removed."})

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsRDCO])
    def decline(self, request, pk=None):
        """
        POST /delete-requests/<id>/decline/
        Marks the request declined, restores the record to published, and notifies the requester.
        """
        from django.utils import timezone
        dr = self.get_object()
        if dr.status != "pending":
            return Response(
                {"detail": f"Request is already '{dr.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        dr.status      = "declined"
        dr.reviewed_by = request.user
        dr.reviewed_at = timezone.now()
        dr.save(update_fields=["status", "reviewed_by", "reviewed_at"])
        # Restore the record to its pre-deletion visible state
        if dr.previous_pipeline_status:
            dr.record.pipeline_status = dr.previous_pipeline_status
        else:
            rt = dr.record.record_type.name if dr.record.record_type else ""
            dr.record.pipeline_status = "approved" if rt == "Proposal" else "published"
        dr.record.save(update_fields=["pipeline_status", "updated_at"])
        notify_delete_declined(dr, reviewed_by=request.user)
        return Response({"detail": "Delete request declined. The record has been restored."})
